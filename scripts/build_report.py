"""
build_report.py
----------------
monitor_ollama.py が出力した results/Functional Testing/summary.csv (4モデル分の実測ログ)を読み込み、
同じ結果フォルダに results/Functional Testing/summary_report.xlsx を生成する。

CSV自体は「1シートのみ」のフォーマットなので複数シート化はできない。
そのため、summary.csv はログとしてそのまま残しつつ、
Excelブック(.xlsx)を新規生成して以下の構成にする。

  1. "全実行ログ"      … summary.csv の中身をそのまま transcribe(生ログ)
  2. "モデル比較(平均)" … 4モデルを横並びで比較する表。色分けで
                         「応答速度が速い/遅い」「VRAM見積もりからの乖離」
                         「実行ごとのブレが大きい(不安定)」を可視化する
  3. "所見"            … 自動検出した注意点のテキストまとめ

使い方:
    python build_report.py
    (results/summary.csv を読み、results/summary_report.xlsx に出力)
"""

from __future__ import annotations

import csv
import statistics
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
RESULTS_DIR = SCRIPT_DIR / "results" / "Functional Testing"
SUMMARY_CSV = RESULTS_DIR / "summary.csv"
REPORT_XLSX = RESULTS_DIR / "summary_report.xlsx"

# 設計時見積もり(1日目設計とモデル選定 / 2日目ノートの記録表より)
DESIGN_VRAM_MIB = {
    "gemma4:26b": 16 * 1024,
    "gpt-oss:20b": 16 * 1024,
    "devstral-small-2:24b": 14 * 1024,
    "phi4-mini:latest": int(2.5 * 1024),
}
MODEL_DISPLAY_NAME = {
    "gemma4:26b": "Gemma4 26B-A4B(計画)",
    "gpt-oss:20b": "gpt-oss-20b(速度)",
    "devstral-small-2:24b": "Devstral Small 2 24B(コーディング)",
    "phi4-mini:latest": "Phi-4-mini(ルーター)",
}
MODEL_ORDER = list(DESIGN_VRAM_MIB.keys())

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ISSUE_FILL = PatternFill("solid", fgColor="F4B6B6")  # 問題あり(赤系)
CAUTION_FILL = PatternFill("solid", fgColor="FCE4A6")  # 要注意(黄系)
GOOD_FILL = PatternFill("solid", fgColor="C6E7B0")  # 良好(緑系)
BOLD = Font(bold=True)


def to_float(v):
    if v in (None, "", "None"):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def load_rows():
    with open(SUMMARY_CSV, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def group_by_model(rows):
    """CSVには同一モデルを複数回測定し直した履歴が残ることがある
    (例: phi4を3回試行→異常値に気づいて6回で再測定、など)。
    同じモデルについて「直近の avg 行」で区切られる直前の連続した
    run行の集まり(=最新の測定バッチ)だけを採用し、古い測定は
    参考データとして混ぜ込まない。
    """
    grouped: dict[str, dict[str, list[dict]]] = {}
    pending_runs: dict[str, list[dict]] = {}
    for row in rows:
        model = row["model"]
        label = row["label"]
        grouped.setdefault(model, {"runs": [], "avg": None})
        pending_runs.setdefault(model, [])
        if "_avg(" in label:
            grouped[model]["avg"] = row
            grouped[model]["runs"] = pending_runs[model]  # 直近バッチで上書き
            pending_runs[model] = []
        else:
            pending_runs[model].append(row)
    # avg行を伴わない末尾の試行(measurement中断など)しか無いモデルは
    # そのまま runs として使う
    for model, runs in pending_runs.items():
        if runs and not grouped[model]["runs"]:
            grouped[model]["runs"] = runs
    return grouped


def compute_model_stats(model: str, data: dict) -> dict:
    runs = [r for r in data["runs"] if "_avg(" not in r["label"]]
    runs_sorted = runs  # 実行順(CSV追記順)のまま = run1, run2, run3...

    vram_peaks = [to_float(r["vram_used_peak_mib"]) for r in runs_sorted]
    vram_peaks = [v for v in vram_peaks if v is not None]
    gpu_util_avgs = [to_float(r["gpu_util_avg_pct"]) for r in runs_sorted]
    gpu_util_avgs = [v for v in gpu_util_avgs if v is not None]
    eval_rates = [to_float(r["eval_rate_tps"]) for r in runs_sorted]
    eval_rates = [v for v in eval_rates if v is not None]
    ttfts = [to_float(r["ttft_ms"]) for r in runs_sorted]
    ttfts = [v for v in ttfts if v is not None]

    cold_ttft = ttfts[0] if ttfts else None
    warm_ttfts = ttfts[1:] if len(ttfts) > 1 else []
    warm_ttft_avg = statistics.mean(warm_ttfts) if warm_ttfts else None

    vram_peak_max = max(vram_peaks) if vram_peaks else None
    vram_variation = (
        (max(vram_peaks) - min(vram_peaks)) if len(vram_peaks) > 1 else 0
    )

    design_vram = DESIGN_VRAM_MIB.get(model)
    vram_deviation_pct = (
        round((vram_peak_max - design_vram) / design_vram * 100, 1)
        if vram_peak_max is not None and design_vram
        else None
    )

    return {
        "model": model,
        "display": MODEL_DISPLAY_NAME.get(model, model),
        "n_runs": len(runs_sorted),
        "vram_peak_max": vram_peak_max,
        "vram_variation": vram_variation,
        "design_vram": design_vram,
        "vram_deviation_pct": vram_deviation_pct,
        "gpu_util_avg": round(statistics.mean(gpu_util_avgs), 1) if gpu_util_avgs else None,
        "eval_rate_avg": round(statistics.mean(eval_rates), 2) if eval_rates else None,
        "cold_ttft_ms": round(cold_ttft, 1) if cold_ttft is not None else None,
        "warm_ttft_ms": round(warm_ttft_avg, 1) if warm_ttft_avg is not None else None,
    }


def build_findings(stats_list: list[dict]) -> list[str]:
    findings = []
    # 生成速度の比較
    fastest = max(stats_list, key=lambda s: s["eval_rate_avg"] or 0)
    slowest = min(stats_list, key=lambda s: s["eval_rate_avg"] or float("inf"))
    findings.append(
        f"【生成速度】最速は {fastest['display']}({fastest['eval_rate_avg']} tokens/s)、"
        f"最遅は {slowest['display']}({slowest['eval_rate_avg']} tokens/s)。"
    )
    if slowest["eval_rate_avg"] and fastest["eval_rate_avg"]:
        ratio = fastest["eval_rate_avg"] / slowest["eval_rate_avg"]
        if ratio >= 3:
            findings.append(
                f"  → {slowest['display']} は最速モデルの約{ratio:.1f}分の1の生成速度。"
                f"コーディング支援用途でテンポが遅く感じられないか要確認。"
            )

    # VRAM見積もりからの乖離
    for s in stats_list:
        if s["vram_deviation_pct"] is None:
            continue
        if abs(s["vram_deviation_pct"]) >= 50:
            findings.append(
                f"【VRAM乖離・要注意】{s['display']}: 実測ピーク {s['vram_peak_max']:.0f}MiB は"
                f"設計見積もり {s['design_vram']}MiB から {s['vram_deviation_pct']:+.1f}% 乖離。"
            )
        elif abs(s["vram_deviation_pct"]) >= 15:
            findings.append(
                f"【VRAM乖離・軽微】{s['display']}: 実測ピーク {s['vram_peak_max']:.0f}MiB は"
                f"設計見積もり {s['design_vram']}MiB から {s['vram_deviation_pct']:+.1f}% 乖離。"
            )

    # 試行間のVRAMブレ(不安定判定)
    for s in stats_list:
        if s["vram_variation"] and s["vram_variation"] >= 5000:
            findings.append(
                f"【不安定・要確認】{s['display']}: 試行間でVRAM使用量が {s['vram_variation']:.0f}MiB "
                f"もブレている(初回だけ大きく異なる等)。ロード状態が試行ごとに変わっている可能性。"
            )

    # コールド/ウォームTTFTの差
    for s in stats_list:
        if s["cold_ttft_ms"] and s["warm_ttft_ms"] and s["warm_ttft_ms"] > 0:
            ratio = s["cold_ttft_ms"] / s["warm_ttft_ms"]
            if ratio >= 10:
                findings.append(
                    f"【初回ロード遅延】{s['display']}: 初回TTFT {s['cold_ttft_ms']:.0f}ms は"
                    f"ウォーム時平均 {s['warm_ttft_ms']:.0f}ms の約{ratio:.0f}倍。"
                    f"外付けHDD/ディスク読み込みが支配的な可能性(常駐運用なら影響小)。"
                )

    return findings


def write_raw_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = "全実行ログ"
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        values = []
        for h in headers:
            v = row[h]
            fv = to_float(v)
            values.append(fv if fv is not None else v)
        ws.append(values)
        # 平均行を薄く強調
        if "_avg(" in row["label"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).font = BOLD
    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(h) + 2)
    ws.freeze_panes = "A2"


def write_comparison_sheet(wb: Workbook, stats_list: list[dict]):
    ws = wb.create_sheet("モデル比較(平均)")
    headers = [
        "モデル",
        "VRAM実測ピーク(MiB)",
        "VRAM設計見積もり(MiB)",
        "VRAM乖離(%)",
        "試行間VRAMブレ(MiB)",
        "GPU使用率 平均(%)",
        "初回TTFT(ms・コールド)",
        "定常TTFT(ms・ウォーム平均)",
        "生成速度 平均(tokens/s)",
        "評価コメント",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 32

    start_row = 2
    for s in stats_list:
        comment_parts = []
        if s["vram_deviation_pct"] is not None and abs(s["vram_deviation_pct"]) >= 50:
            comment_parts.append("VRAM見積もり大幅乖離")
        if s["vram_variation"] and s["vram_variation"] >= 5000:
            comment_parts.append("試行間の挙動が不安定")
        comment = " / ".join(comment_parts) if comment_parts else "見積もり通り・安定"
        ws.append(
            [
                s["display"],
                s["vram_peak_max"],
                s["design_vram"],
                s["vram_deviation_pct"],
                round(s["vram_variation"], 1) if s["vram_variation"] else 0,
                s["gpu_util_avg"],
                s["cold_ttft_ms"],
                s["warm_ttft_ms"],
                s["eval_rate_avg"],
                comment,
            ]
        )
    end_row = start_row + len(stats_list) - 1

    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(16, len(h) + 2)

    # 色分け1: 生成速度(tokens/s) 高いほど緑・低いほど赤 = 速いモデルが一目でわかる
    speed_col = "I"
    ws.conditional_formatting.add(
        f"{speed_col}{start_row}:{speed_col}{end_row}",
        ColorScaleRule(
            start_type="min", start_color="F4B6B6",
            mid_type="percentile", mid_value=50, mid_color="FCE4A6",
            end_type="max", end_color="C6E7B0",
        ),
    )
    # 色分け2: 定常TTFT 低いほど緑(速い)・高いほど赤(遅い)
    ttft_col = "H"
    ws.conditional_formatting.add(
        f"{ttft_col}{start_row}:{ttft_col}{end_row}",
        ColorScaleRule(
            start_type="min", start_color="C6E7B0",
            mid_type="percentile", mid_value=50, mid_color="FCE4A6",
            end_type="max", end_color="F4B6B6",
        ),
    )
    # 色分け3: VRAM乖離(%) 絶対値が大きいほど赤(設計から外れている=問題の芽)
    dev_col = "D"
    for r in range(start_row, end_row + 1):
        cell = ws[f"{dev_col}{r}"]
        val = cell.value
        if val is None:
            continue
        if abs(val) >= 50:
            cell.fill = ISSUE_FILL
        elif abs(val) >= 15:
            cell.fill = CAUTION_FILL
        else:
            cell.fill = GOOD_FILL

    # 色分け4: 試行間VRAMブレが大きい(不安定)行を赤で強調
    var_col = "E"
    for r in range(start_row, end_row + 1):
        cell = ws[f"{var_col}{r}"]
        val = cell.value
        if val is None:
            continue
        if val >= 5000:
            cell.fill = ISSUE_FILL
        elif val >= 1000:
            cell.fill = CAUTION_FILL
        else:
            cell.fill = GOOD_FILL

    # 評価コメント列: 問題ありなら赤字太字
    comment_col = "J"
    for r in range(start_row, end_row + 1):
        cell = ws[f"{comment_col}{r}"]
        if cell.value and cell.value != "見積もり通り・安定":
            cell.font = Font(color="C0392B", bold=True)
        else:
            cell.font = Font(color="2E7D32")

    ws.freeze_panes = "A2"

    # 凡例
    legend_row = end_row + 3
    ws.cell(row=legend_row, column=1, value="凡例:").font = BOLD
    ws.cell(row=legend_row + 1, column=1, value="緑=良好/速い").fill = GOOD_FILL
    ws.cell(row=legend_row + 2, column=1, value="黄=軽微な乖離/要観察").fill = CAUTION_FILL
    ws.cell(row=legend_row + 3, column=1, value="赤=設計乖離大/不安定=問題の可能性").fill = ISSUE_FILL


def write_findings_sheet(wb: Workbook, findings: list[str]):
    ws = wb.create_sheet("所見")
    ws.cell(row=1, column=1, value="自動検出した所見(2日目モデル動作検証より)").font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 110
    r = 3
    if not findings:
        ws.cell(row=r, column=1, value="特筆すべき乖離・不安定な挙動は検出されませんでした。")
        return
    for line in findings:
        cell = ws.cell(row=r, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True)
        if "要注意" in line or "不安定" in line:
            cell.fill = ISSUE_FILL
        elif "軽微" in line or "初回ロード遅延" in line:
            cell.fill = CAUTION_FILL
        r += 1


def main():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            try:
                stream.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass

    rows = load_rows()
    if not rows:
        print("[warn] summary.csv にデータがありません")
        return

    grouped = group_by_model(rows)
    stats_list = []
    for model in MODEL_ORDER:
        if model not in grouped:
            continue
        stats_list.append(compute_model_stats(model, grouped[model]))
    # CSVに存在するがMODEL_ORDER未登録のモデルも末尾に追加
    for model in grouped:
        if model not in MODEL_ORDER:
            stats_list.append(compute_model_stats(model, grouped[model]))

    findings = build_findings(stats_list)

    wb = Workbook()
    write_raw_sheet(wb, rows)
    write_comparison_sheet(wb, stats_list)
    write_findings_sheet(wb, findings)

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(REPORT_XLSX)
    print(f"[info] レポートを出力しました: {REPORT_XLSX}")
    print(f"[info] 検出された所見: {len(findings)}件")
    for f in findings:
        print(f"  - {f}")


if __name__ == "__main__":
    main()
